# reporte.py
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io
from datetime import date

def crear_reporte_excel(
    direccion: str,
    total_pendientes: int,
    tabla_cursos: list[list],
    tabla_areas: list[list],
    fecha_texto: str | None = None
) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard Dirección"

    bold_font = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    purple_fill = PatternFill("solid", fgColor="800080")
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('A1:D1')
    ws['A1'] = "CURSOS PENDIENTES POR DIRECCIÓN"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center

    ws.merge_cells('A2:D2')
    ws['A2'] = f"DIRECCIÓN: {direccion}"
    ws['A2'].font = Font(bold=True)
    ws['A2'].alignment = center

    ws['F1'] = fecha_texto or date.today().strftime("%d/%m/%Y")
    ws['F1'].alignment = Alignment(horizontal="right")
    ws['F2'] = "Total pendientes"
    ws['F3'] = int(total_pendientes)
    ws['F3'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['F3'].fill = PatternFill("solid", fgColor="FF4D6D")
    ws['F3'].alignment = center

    start_row = 5
    for i, row in enumerate(tabla_cursos):
        for j, value in enumerate(row):
            cell = ws.cell(row=start_row + i, column=j + 1, value=value)
            if i == 0:
                cell.font = header_font
                cell.fill = purple_fill
            cell.alignment = center

    start_row_area = 12
    ws.cell(row=start_row_area, column=1, value="Cursos Pendientes por Área").font = bold_font

    for i, row in enumerate(tabla_areas):
        for j, value in enumerate(row):
            cell = ws.cell(row=start_row_area + 1 + i, column=j + 1, value=value)
            if i == 0:
                cell.font = header_font
                cell.fill = purple_fill
            cell.alignment = center

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["F"].width = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
